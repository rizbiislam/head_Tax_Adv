import argparse
from pathlib import Path
import segregate

try:
    import pandas as pd
except Exception:
    pd = None

from openpyxl import load_workbook


def scan_with_openpyxl(path, total_rate_names):
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [h for h in rows[0]]
    # find total rate column index
    total_idx = None
    for name in total_rate_names:
        if name in headers:
            total_idx = headers.index(name)
            break
    primary_idx = None
    # try common primary key names
    for pk in ['PrimaryKey', 'Id', 'HoldingId']:
        if pk in headers:
            primary_idx = headers.index(pk)
            break

    found = []
    for r in rows[1:]:
        total_val = r[total_idx] if total_idx is not None else None
        try:
            total_num = float(total_val) if total_val is not None else None
        except Exception:
            total_num = None
        if total_num is not None and abs(total_num - 7.0) < 1e-6:
            pk = r[primary_idx] if primary_idx is not None else None
            found.append({'primary': pk, 'total': total_num})
    return found


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', required=True)
    ap.add_argument('--config-dir', default=str(Path(__file__).parent / 'config'))
    args = ap.parse_args()

    # try to load config
    try:
        heads, col_map, overrides = segregate.load_config(Path(args.config_dir))
    except Exception as e:
        print('Failed to load config:', e)
        heads, col_map, overrides = [], {'master_table': {}}, {}

    total_rate_names = ['TotalTaxRates', 'TotalTaxRate', 'total_rate_check']

    if pd is not None:
        master_df = pd.read_excel(args.master)
        m = col_map.get('master_table', {})
        total_rate_col = segregate._resolve_column(master_df, m, 'total_rate_check', 'TotalTaxRates', 'TotalTaxRate', default='TotalTaxRates')
        if total_rate_col and total_rate_col in master_df.columns:
            master_df['_total_check'] = master_df[total_rate_col].apply(lambda v: segregate._coerce_numeric(v))
        else:
            master_df['_total_check'] = None

        target = 7.0
        mask = master_df['_total_check'].apply(lambda v: v is not None and abs(v - target) < 1e-6)
        sevens = master_df[mask]
        print(f"Found {len(sevens)} holdings with TotalTaxRates == {target} (column: {total_rate_col})")
        if len(sevens):
            pk = col_map.get('master_table', {}).get('primary_key', 'Id')
            cols = [c for c in [total_rate_col, pk] if c in master_df.columns]
            print(sevens[cols].head(50).to_string(index=False))

        try:
            out_df, exc_df = segregate.segregate(master_df, heads, col_map, iuser=360, idate='2026-08-12', overrides=overrides)
            print(f"Segregate generated {len(out_df)} rows and flagged {len(exc_df)} exceptions")
            if len(exc_df):
                print('\nExceptions sample:')
                print(exc_df.head(50).to_string(index=False))
        except Exception as e:
            print('Segregate run failed (pandas present but error):', e)
    else:
        print('pandas is not importable in this environment — falling back to openpyxl-only scan')
        found = scan_with_openpyxl(args.master, total_rate_names)
        print(f"Found {len(found)} holdings with TotalTaxRates == 7.0 (openpyxl scan)")
        if found:
            for r in found[:50]:
                print(r)
        print('\nTo run full segregation you need pandas/numpy installed. Install via:')
        print('    pip install -r requirements.txt')


if __name__ == '__main__':
    main()
